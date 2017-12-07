import openpyxl
import csv
import time
from dicts import *
from fix_groupon import *
import datetime
import sys
import pymysql

def process_output(groupon_file, commerce_file, staples_file):
	# connect to db for SKUs and UPCs
	conn = pymysql.connect(host='127.0.0.1', user='root', passwd='Greengiant90',
							db='mysql', charset='utf8')
	cur = conn.cursor()
	cur.execute("USE star_interactive")

	output_wb = openpyxl.Workbook()
	output_sheet = output_wb.active

	# write columns for new output wb, set col. width
	col_width = 20
	format_file = "FORMAT SHEET.xlsx" # Always use this spreadsheet as the template for the column names/formatting
	format_wb = openpyxl.load_workbook(format_file)
	format_sheet = format_wb.active
	for col in range(1, format_sheet.max_column + 1):
		col_letter = openpyxl.cell.cell.get_column_letter(col)
		output_sheet[col_letter + str(1)] = format_sheet[col_letter + str(1)].value
		output_sheet.column_dimensions[col_letter].width = col_width

	final_col = format_sheet.max_column + 1

	# row offset for multiple input sheets
	offset = 0
	# error_rows for final checking
	error_rows = set()

	if commerce_file:
		print('Adding CommerceHub')
		output_sheet, offset, error_rows = process_sheet(commerce_file, final_col, output_sheet, commercehub_dict, offset, cur, error_rows)

	if groupon_file:
		print('Adding Groupon')
		output_sheet, offset, error_rows = process_sheet(groupon_file, final_col, output_sheet, groupon_dict, offset, cur, error_rows)

	if staples_file:
		print('Adding Staples')
		output_sheet, offset, error_rows = process_sheet(staples_file, final_col, output_sheet, staples_dict, offset, cur, error_rows)


	today = datetime.date.today()
	output_file = today.strftime("%m-%d-%Y") + " ORDERS.xlsx"
	output_wb.save(output_file)
	print('Done.')
	print('Total SKUs:', output_sheet.max_row - 1)
	print('-----')
	for row in error_rows:
		print('WARNING! Potential error on row', row)

	cur.close()
	conn.close()